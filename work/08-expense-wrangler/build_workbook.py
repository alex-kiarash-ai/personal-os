"""Builds the branded Expense Tracker workbook with ALL real formulas.
Reusable: the automation re-runs this to regenerate the monthly report from the Expense Log.
Currency: SEK (kr) per brand-config. Change CURFMT to switch.
"""
import sys, json, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

args = [a for a in sys.argv[1:] if not a.startswith("--")]
OUT = args[0] if args else "expense-tracker.xlsx"
EMPTY = "--empty" in sys.argv  # deliver a clean tracker (no example rows)
# --data <path>: load accumulated rows from JSON {"rows": [[date,vendor,amount,cat,tax,status,note],...]}
DATA = None
for i, a in enumerate(sys.argv):
    if a == "--data" and i + 1 < len(sys.argv):
        DATA = sys.argv[i + 1]

TEAL = "005F73"; CYAN = "0A9396"; ORANGE = "EE9B00"; TINT = "FFF5E1"; WHITE = "FFFFFF"
CURFMT = '#,##0.00 "kr"'; PCTFMT = '0.0%'
hdr_fill = PatternFill("solid", fgColor=TEAL)
hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
title_font = Font(name="Calibri", bold=True, color=TEAL, size=15)
bold_teal = Font(name="Calibri", bold=True, color=TEAL)
alt_fill = PatternFill("solid", fgColor=TINT)
thin = Side(style="thin", color="D9DEE8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center")

CATS = ["Meals", "Travel", "Software", "Office", "Subscriptions", "Other"]
# Example seed rows (clearly marked; cleared on first real use). Prove the formulas compute.
SEED = [] if EMPTY else [
    ("2026-04-03", "Sample Cafe",   450.00, "Meals", True,  "Verified", "Sample meal"),
    ("2026-04-14", "Hetzner",        110.00, "Software", True, "Verified", "VPS monthly"),
    ("2026-05-02", "SJ",             390.00, "Travel", True,  "Verified", "Train Sthlm"),
    ("2026-05-19", "Anthropic",      220.00, "Software", True, "Verified", "Claude API"),
    ("2026-06-01", "WeWork",         180.00, "Office", True,  "Unmatched", "Day pass"),
]
# Real accumulated data takes precedence over example SEED when --data is given.
if DATA:
    with open(DATA, encoding="utf-8") as fh:
        SEED = [tuple(r) for r in json.load(fh)["rows"]]

wb = openpyxl.Workbook()

# ---------- Sheet 1: Expense Log ----------
log = wb.active; log.title = "Expense Log"
headers = ["Date", "Vendor", "Amount", "Category", "Tax Deductible", "Status", "Month", "Quarter"]
log.append(headers)
for c in range(1, len(headers)+1):
    cell = log.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border; cell.alignment = center
DATA_START = 2
for i, (d, v, a, cat, tax, st, note) in enumerate(SEED):
    r = DATA_START + i
    log.cell(r, 1, d)
    log.cell(r, 2, v)
    log.cell(r, 3, a).number_format = CURFMT
    log.cell(r, 4, cat)
    log.cell(r, 5, "Yes" if tax else "No")
    log.cell(r, 6, st)
    # Month + Quarter are FORMULAS derived from the Date (real formulas, not typed)
    log.cell(r, 7, f'=IF($A{r}="","",TEXT(DATEVALUE($A{r}),"YYYY-MM"))')
    log.cell(r, 8, f'=IF($A{r}="","",YEAR(DATEVALUE($A{r}))&"-Q"&ROUNDUP(MONTH(DATEVALUE($A{r}))/3,0))')
    if i % 2 == 1:
        for c in range(1, 9): log.cell(r, c).fill = alt_fill
widths = [12, 22, 14, 14, 14, 12, 10, 10]
for c, w in enumerate(widths, 1): log.column_dimensions[get_column_letter(c)].width = w
log.freeze_panes = "A2"
LASTSEED = DATA_START + len(SEED) - 1

# ---------- Sheet 2: Monthly Summary ----------
ms = wb.create_sheet("Monthly Summary")
ms["A1"] = "Monthly Summary"; ms["A1"].font = title_font
ms.append([])
hrow = 3
ms.cell(hrow, 1, "Month")
for j, cat in enumerate(CATS): ms.cell(hrow, 2+j, cat)
ms.cell(hrow, 2+len(CATS), "Total")
for c in range(1, 3+len(CATS)):
    cell = ms.cell(hrow, c); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border; cell.alignment = center
# months Apr-Dec 2026 as rows (formula-driven SUMIFS over the log columns)
months = [f"2026-{m:02d}" for m in range(1, 13)]
for k, mo in enumerate(months):
    r = hrow + 1 + k
    ms.cell(r, 1, mo)
    for j, cat in enumerate(CATS):
        col = get_column_letter(2+j)
        ms.cell(r, 2+j, f"=SUMIFS('Expense Log'!$C:$C,'Expense Log'!$G:$G,$A{r},'Expense Log'!$D:$D,{col}${hrow})").number_format = CURFMT
    tcol = get_column_letter(2+len(CATS))
    first = get_column_letter(2); last = get_column_letter(1+len(CATS))
    ms.cell(r, 2+len(CATS), f"=SUM({first}{r}:{last}{r})").number_format = CURFMT
# grand total row
gr = hrow + 1 + len(months)
ms.cell(gr, 1, "TOTAL").font = bold_teal
for j in range(len(CATS)+1):
    col = get_column_letter(2+j)
    ms.cell(gr, 2+j, f"=SUM({col}{hrow+1}:{col}{gr-1})").number_format = CURFMT
    ms.cell(gr, 2+j).font = bold_teal
ms.column_dimensions["A"].width = 12
for c in range(2, 3+len(CATS)): ms.column_dimensions[get_column_letter(c)].width = 14

# ---------- Sheet 3: Quarterly Summary ----------
qs = wb.create_sheet("Quarterly Summary")
qs["A1"] = "Quarterly Summary"; qs["A1"].font = title_font
qs.append([])
qh = 3
for c, name in enumerate(["Quarter", "Total", "QoQ Change %"], 1):
    cell = qs.cell(qh, c, name); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border; cell.alignment = center
quarters = [f"2026-Q{q}" for q in range(1, 5)]
for k, q in enumerate(quarters):
    r = qh + 1 + k
    qs.cell(r, 1, q)
    qs.cell(r, 2, f"=SUMIFS('Expense Log'!$C:$C,'Expense Log'!$H:$H,$A{r})").number_format = CURFMT
    if k == 0:
        qs.cell(r, 3, "=\"\"")  # no prior quarter
    else:
        qs.cell(r, 3, f'=IF(B{r-1}=0,"",(B{r}-B{r-1})/B{r-1})').number_format = PCTFMT
qs.column_dimensions["A"].width = 12; qs.column_dimensions["B"].width = 14; qs.column_dimensions["C"].width = 14

# ---------- Sheet 4: Category Breakdown ----------
cb = wb.create_sheet("Category Breakdown")
cb["A1"] = "Category Breakdown"; cb["A1"].font = title_font
cb.append([])
ch = 3
for c, name in enumerate(["Category", "Total", "% of Total", "Avg / Month"], 1):
    cell = cb.cell(ch, c, name); cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border; cell.alignment = center
for k, cat in enumerate(CATS):
    r = ch + 1 + k
    cb.cell(r, 1, cat)
    cb.cell(r, 2, f"=SUMIF('Expense Log'!$D:$D,$A{r},'Expense Log'!$C:$C)").number_format = CURFMT
    # distinct months with any spend = count of nonblank Month cells (whole-column); avg = total / that count
    cb.cell(r, 4, f'=IF($G${ch}=0,0,B{r}/$G${ch})').number_format = CURFMT
grow = ch + 1 + len(CATS)
cb.cell(grow, 1, "TOTAL").font = bold_teal
cb.cell(grow, 2, f"=SUM(B{ch+1}:B{grow-1})").number_format = CURFMT; cb.cell(grow, 2).font = bold_teal
# % of total references the grand total
for k in range(len(CATS)):
    r = ch + 1 + k
    cb.cell(r, 3, f'=IF($B${grow}=0,0,B{r}/$B${grow})').number_format = PCTFMT
cb.cell(grow, 3, f'=IF($B${grow}=0,0,SUM(C{ch+1}:C{grow-1}))').number_format = PCTFMT; cb.cell(grow,3).font = bold_teal
# helper: distinct month count in G3 (used by Avg/Month)
cb.cell(ch, 7, "distinct months").font = Font(size=8, color="9CA3AF")
cb.cell(ch, 7).value = '=SUMPRODUCT((\'Expense Log\'!G2:G100000<>"")/COUNTIF(\'Expense Log\'!G2:G100000,\'Expense Log\'!G2:G100000&""))'
for c, w in zip("ABCD", [16, 14, 12, 14]): cb.column_dimensions[c].width = w

# ---------- Uncaptured Burn row in Monthly Summary (P6 rider, first appears at 07-31 close) ----------
# Monthly burn target from vault/projects/runway/status.md (monthly_burn_sek = 24000)
MONTHLY_BURN_TARGET = 24000
july_row_in_ms = hrow + 1 + 6          # 2026-07 is month index 6 (Jan=0)
total_col_letter = get_column_letter(2 + len(CATS))  # H = the Total column
ub_label_row = gr + 1
ub_row = gr + 2
ms.cell(ub_label_row, 1, "--- Jul 2026 close ---").font = Font(name="Calibri", color="9CA3AF", size=9)
ms.cell(ub_row, 1, "Uncaptured burn est. (2026-07)").font = Font(name="Calibri", italic=True, color=TEAL, size=10)
ms.cell(ub_row, 2 + len(CATS),
        f"=MAX(0,{MONTHLY_BURN_TARGET}-{total_col_letter}{july_row_in_ms})").number_format = CURFMT

# ---------- Sheet 5: Alex Costs (P9a merge, 2026-07-31) ----------
# Receipt-backed rows from the 2026-07-02 alex-cost-tracker.xlsx, merged here at the monthly close.
# Totals use real formulas; data rows are receipt-backed inputs (not computed values).
ac = wb.create_sheet("Alex Costs")
ac["A1"] = "Alex Costs"; ac["A1"].font = title_font
ac.append([])

# FX rate helper cells (editable inputs, referenced by run-rate formulas)
ac["A3"] = "EUR/SEK (approx)";      ac["B3"] = 11.20; ac["B3"].number_format = "0.00"
ac["A4"] = "USD/SEK (approx)";      ac["B4"] = 9.60;  ac["B4"].number_format = "0.00"
ac["A5"] = "Hetzner monthly (EUR)"; ac["B5"] = 9.99;  ac["B5"].number_format = "0.00"
ac.append([])

# Ledger — receipt-backed rows (merged from alex-cost-tracker.xlsx, 2026-07-31)
ACL_START = 8
for c_idx, name in enumerate(["Date", "Type", "Vendor", "Description", "Amount (SEK)"], 1):
    cell = ac.cell(ACL_START - 1, c_idx, name)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.border = border; cell.alignment = center
AC_ROWS = [
    ("2026-04-24", "Annual", "Anthropic",  "Claude Pro annual €225 @ EUR/SEK 11.20",   2520.00),
    ("2026-05-16", "API",    "Anthropic",  "API top-up #1 $25 @ USD/SEK 9.60",          240.00),
    ("2026-06-03", "Infra",  "Hetzner",    "VPS May invoice €7.44 @ EUR/SEK 11.20",      83.33),
    ("2026-06-22", "API",    "Anthropic",  "API top-up #2 $25 @ USD/SEK 9.60",          240.00),
    ("2026-07-02", "API",    "Anthropic",  "API top-up #3 $25 @ USD/SEK 9.60",          240.00),
    ("2026-07-03", "Domain", "Cloudflare", "Domain renewal $10.46 @ USD/SEK 9.60",      100.42),
]
ACL_END = ACL_START + len(AC_ROWS) - 1
for i, (d, tp, vendor, desc, amt) in enumerate(AC_ROWS):
    r = ACL_START + i
    ac.cell(r, 1, d); ac.cell(r, 2, tp); ac.cell(r, 3, vendor); ac.cell(r, 4, desc)
    ac.cell(r, 5, amt).number_format = CURFMT
    if i % 2 == 1:
        for c in range(1, 6): ac.cell(r, c).fill = alt_fill

# Cash Summary (SUMIFS on the Ledger)
sr = ACL_END + 2
ac.cell(sr, 1, "Cash Summary").font = bold_teal
summary_items = [
    ("Total Cash Spent (kr)",       f"=SUM(E{ACL_START}:E{ACL_END})"),
    ("API Top-ups",                 f'=SUMIF(B{ACL_START}:B{ACL_END},"API",E{ACL_START}:E{ACL_END})'),
    ("Infrastructure (Hetzner)",    f'=SUMIF(C{ACL_START}:C{ACL_END},"Hetzner",E{ACL_START}:E{ACL_END})'),
    ("Annual Plan (Anthropic Pro)", f'=SUMIF(B{ACL_START}:B{ACL_END},"Annual",E{ACL_START}:E{ACL_END})'),
    ("Domain / Other",              f'=SUMIF(B{ACL_START}:B{ACL_END},"Domain",E{ACL_START}:E{ACL_END})'),
]
for k, (label, formula) in enumerate(summary_items):
    r = sr + 1 + k
    ac.cell(r, 1, label)
    ac.cell(r, 2, formula).number_format = CURFMT
    if k == 0: ac.cell(r, 1).font = Font(name="Calibri", bold=True, color=TEAL)

# Monthly Run Rate (Accrual)
rr_start = sr + 1 + len(summary_items) + 1
ac.cell(rr_start, 1, "Monthly Run Rate (Accrual Model)").font = bold_teal
rr_items = [
    ("Claude Pro / month",        f'=SUMIF(B{ACL_START}:B{ACL_END},"Annual",E{ACL_START}:E{ACL_END})/12'),
    ("Hetzner / month",           "=$B$3*$B$5"),        # EUR/SEK rate * €9.99
    ("Cloudflare / month",        f'=SUMIF(B{ACL_START}:B{ACL_END},"Domain",E{ACL_START}:E{ACL_END})/12'),
]
for k, (label, formula) in enumerate(rr_items):
    r = rr_start + 1 + k
    ac.cell(r, 1, label); ac.cell(r, 2, formula).number_format = CURFMT
fixed_row = rr_start + 1 + len(rr_items)
ac.cell(fixed_row, 1, "Fixed total / month").font = Font(name="Calibri", bold=True, color=TEAL)
ac.cell(fixed_row, 2, f"=SUM(B{rr_start+1}:B{fixed_row-1})").number_format = CURFMT
# API projection: 3 top-ups over ~2.5 months of actual data (Apr 24 → Jul 2)
api_row = fixed_row + 1
ac.cell(api_row, 1, "API projected / month (3 top-ups / 2.5 mo)")
ac.cell(api_row, 2, f'=SUMIF(B{ACL_START}:B{ACL_END},"API",E{ACL_START}:E{ACL_END})/2.5').number_format = CURFMT
total_rr_row = api_row + 1
ac.cell(total_rr_row, 1, "Est. Total Run Rate / month").font = Font(name="Calibri", bold=True, color=TEAL)
ac.cell(total_rr_row, 2, f"=B{fixed_row}+B{api_row}").number_format = CURFMT

# Notes
nr = total_rr_row + 2
ac.cell(nr, 1, "Notes").font = bold_teal
for row_offset, note in enumerate([
    "Merged into #08 Expense Wrangler at 2026-07-31 close (P9a / D10).",
    "July Hetzner invoice pending Gmail scan (est. EUR 9.99 + VAT; update E10 when confirmed).",
    "FX rates approximate; update B3:B4 from card statements when available.",
    "Power BI dashboard: outputs/alex-costs/2026-07-03/alex-cost-dashboard.pbip (open in Desktop).",
]):
    ac.cell(nr + 1 + row_offset, 1, note).font = Font(name="Calibri", color="4A5A5E", size=9)

ac.column_dimensions["A"].width = 42
ac.column_dimensions["B"].width = 14
ac.column_dimensions["C"].width = 14
ac.column_dimensions["D"].width = 50
ac.column_dimensions["E"].width = 14

wb.save(OUT)
print("WROTE", OUT, "| sheets:", wb.sheetnames, "| seed rows:", len(SEED))
