#!/usr/bin/env python
"""
ingest_airbnb.py  -  Parse the harvested Airbnb data into the income model.

Primary source: raw/reservations-*.json (DOM scrape of the host reservations table).
Each scraped row is tab/pipe delimited:
  Status |  Guest | N adults |  CheckIn  CheckOut  Booked | time |  Listing  ConfCode  Payout  | Details

Builds a fresh, correctly-sized workbook (all real formulas) at
  outputs/airbnb-host/<today>/airbnb-studio-income-model.xlsx
and a clean raw/bookings-normalized.json for the Notion sync.

No network, no secrets.
"""
import sys, re, json, glob, datetime as dt
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
RAW = HERE / "raw"
ROOT = HERE.parent.parent
TODAY = dt.date.today().isoformat()
OUT_DIR = ROOT / "outputs" / "airbnb-host" / TODAY
OUT_DIR.mkdir(parents=True, exist_ok=True)
MODEL = OUT_DIR / "airbnb-studio-income-model.xlsx"

STATUS_MAP = {"confirmed": "Confirmed", "trip requested": "Pending", "past guest": "Completed"}

# ---- brand ----
TEAL="005F73"; CYAN="0A9396"; AQUA="94D2BD"; INK="001219"; TINT="FFF5E1"; WHITE="FFFFFF"
SEK='#,##0 "kr"'; PCT='0.0%'; DATEF='yyyy-mm-dd'


def newest(pattern):
    files = sorted(glob.glob(str(RAW / pattern)), reverse=True)
    return Path(files[0]) if files else None


def parse_money(s):
    s = (s or "").strip()
    if not s or "kr" not in s.lower():
        return None
    neg = s.lstrip().startswith("-")
    num = re.sub(r"[^\d.]", "", s.replace(",", ""))
    if num == "":
        return None
    val = float(num)
    return -val if neg else val


def parse_date(s):
    s = (s or "").strip()
    for fmt in ("%b %d, %Y", "%B %d, %Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def parse_reservations(path):
    data = json.loads(path.read_text(encoding="utf-8"))
    rows = []
    for raw in data.get("rows", []):
        if "Confirmation Code" in raw and "Status" in raw:
            continue  # header
        parts = [p.strip() for p in raw.split("\t")]
        if len(parts) < 8:
            continue
        status_raw = parts[0].replace("|", "").strip()
        guest = parts[1].split("|")[0].strip()
        ci = parse_date(parts[2])
        co = parse_date(parts[3])
        conf = parts[6].strip()
        payout = parse_money(parts[7])
        if ci is None:
            continue
        if status_raw.lower().startswith("canceled") or status_raw.lower().startswith("cancelled"):
            status = "Cancelled"
        else:
            status = STATUS_MAP.get(status_raw.lower(), "Confirmed")
        nights = (co - ci).days if (ci and co) else None
        rows.append({
            "guest": guest, "check_in": ci.isoformat(),
            "check_out": co.isoformat() if co else None,
            "nights": nights, "status": status, "payout": payout,
            "confirmation": conf, "status_raw": status_raw,
        })
    # newest check-in first
    rows.sort(key=lambda r: r["check_in"], reverse=True)
    return rows


# ---------- workbook builder ----------
def build(rows):
    hdr_fill=PatternFill("solid",fgColor=TEAL); hdr_font=Font(bold=True,color=WHITE)
    title_font=Font(bold=True,color=TEAL,size=16); sub=Font(italic=True,color=CYAN,size=10)
    lbl=Font(bold=True,color=TEAL); val=Font(color=TEAL); tint=PatternFill("solid",fgColor=TINT)
    inp=PatternFill("solid",fgColor="FFF7E6")
    thin=Side(style="thin",color="D8DEEC"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr=Alignment(horizontal="center",vertical="center")
    wb=Workbook()

    # INPUTS
    ws=wb.active; ws.title="Inputs"; ws.sheet_view.showGridLines=False
    ws["A1"]="Airbnb Studio — Income Model"; ws["A1"].font=title_font
    ws["A2"]=f"Auto-filled from your Airbnb account on {TODAY}. Actual Payout = real money. Fill yellow cells for gross/net estimates."
    ws["A2"].font=sub
    ws["A4"]="Assumptions"; ws["A4"].font=lbl
    for i,(name,default,note) in enumerate([
        ("Nightly rate (SEK)",0,"optional — only for the Gross/Est.Net estimate columns"),
        ("Cleaning fee charged to guest (SEK)",0,"optional estimate input"),
        ("Airbnb host service fee (%)",0.03,"~3%, already net in Actual Payout"),
        ("Your cleaning cost (SEK/turnover)",0,"your real cost per turnover"),
        ("Monthly fixed cost (SEK)",0,"rent/utilities share, optional"),
    ]):
        r=5+i
        ws.cell(row=r,column=1,value=name).font=lbl
        c=ws.cell(row=r,column=2,value=default); c.font=val; c.number_format=(PCT if "%" in name else SEK); c.fill=inp; c.border=bd
        ws.cell(row=r,column=3,value=note).font=sub
    ws.column_dimensions["A"].width=42; ws.column_dimensions["B"].width=15; ws.column_dimensions["C"].width=52
    RATE="Inputs!$B$5"; CLEAN="Inputs!$B$6"; SVC="Inputs!$B$7"; CCOST="Inputs!$B$8"

    # BOOKINGS
    ws=wb.create_sheet("Bookings"); ws.sheet_view.showGridLines=False
    ws["A1"]="Bookings Ledger"; ws["A1"].font=title_font
    ws["A2"]=f"{len(rows)} bookings pulled from your account. Actual Payout is real; Gross/Est.Net light up if you set a nightly rate."
    ws["A2"].font=sub
    heads=["Guest","Check-in","Check-out","Nights","Status","Conf. Code","Nightly Rate","Gross","Cleaning Fee","Service Fee","Est. Net","Actual Payout","Month"]
    hrow=4
    for i,h in enumerate(heads,1):
        c=ws.cell(row=hrow,column=i,value=h); c.fill=hdr_fill; c.font=hdr_font; c.alignment=ctr; c.border=bd
    first=hrow+1; last=first+len(rows)-1
    for idx,bk in enumerate(rows):
        r=first+idx
        ws.cell(row=r,column=1,value=bk["guest"])
        try: ws.cell(row=r,column=2,value=dt.date.fromisoformat(bk["check_in"])).number_format=DATEF
        except Exception: pass
        if bk["check_out"]:
            try: ws.cell(row=r,column=3,value=dt.date.fromisoformat(bk["check_out"])).number_format=DATEF
            except Exception: pass
        ws.cell(row=r,column=4,value=f'=IF(AND(B{r}<>"",C{r}<>""),C{r}-B{r},"")')
        ws.cell(row=r,column=5,value=bk["status"])
        ws.cell(row=r,column=6,value=bk["confirmation"])
        ws.cell(row=r,column=7,value=f'=IF(A{r}="","",{RATE})').number_format=SEK
        ws.cell(row=r,column=8,value=f'=IF(OR(D{r}="",D{r}=0,G{r}=0),"",D{r}*G{r})').number_format=SEK
        ws.cell(row=r,column=9,value=f'=IF(A{r}="","",{CLEAN})').number_format=SEK
        ws.cell(row=r,column=10,value=f'=IF(H{r}="","",(H{r}+I{r})*{SVC})').number_format=SEK
        ws.cell(row=r,column=11,value=f'=IF(H{r}="","",H{r}+I{r}-J{r}-{CCOST})').number_format=SEK
        pc=ws.cell(row=r,column=12,value=bk["payout"]); pc.number_format=SEK
        ws.cell(row=r,column=13,value=f'=IF(B{r}="","",TEXT(B{r},"YYYY-MM"))')
        for c in range(1,14):
            cell=ws.cell(row=r,column=c); cell.border=bd
            if idx%2: cell.fill=tint
    widths=[26,12,12,7,13,12,11,11,11,11,11,13,9]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="A5"
    D=f"Bookings!$D${first}:$D${last}"; M=f"Bookings!$M${first}:$M${last}"
    G=f"Bookings!$H${first}:$H${last}"; NET=f"Bookings!$K${first}:$K${last}"
    PAY=f"Bookings!$L${first}:$L${last}"; ST=f"Bookings!$E${first}:$E${last}"

    # MONTHLY SUMMARY
    ws=wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines=False
    ws["A1"]="Monthly Summary"; ws["A1"].font=title_font
    ws["A2"]="Booked nights = Confirmed + Completed. Occupancy = booked nights / days in month."; ws["A2"].font=sub
    mh=["Month","Days","Nights Booked","Occupancy %","Actual Payout","Bookings"]
    hrow=4
    for i,h in enumerate(mh,1):
        c=ws.cell(row=hrow,column=i,value=h); c.fill=hdr_fill; c.font=hdr_font; c.alignment=ctr; c.border=bd
    months=sorted({bk["check_in"][:7] for bk in rows})
    mfirst=hrow+1
    for idx,m in enumerate(months):
        r=mfirst+idx
        ws.cell(row=r,column=1,value=m).font=lbl
        ws.cell(row=r,column=2,value=f'=DAY(EOMONTH(DATEVALUE(A{r}&"-01"),0))')
        ws.cell(row=r,column=3,value=f'=SUMIFS({D},{M},A{r},{ST},"Confirmed")+SUMIFS({D},{M},A{r},{ST},"Completed")')
        ws.cell(row=r,column=4,value=f'=IF(B{r}=0,0,C{r}/B{r})').number_format=PCT
        ws.cell(row=r,column=5,value=f'=SUMIFS({PAY},{M},A{r})').number_format=SEK
        ws.cell(row=r,column=6,value=f'=COUNTIFS({M},A{r})')
        for c in range(1,7):
            cell=ws.cell(row=r,column=c); cell.border=bd
            if idx%2: cell.fill=tint
    tr=mfirst+len(months)
    ws.cell(row=tr,column=1,value="TOTAL").font=Font(bold=True,color=WHITE)
    for c in range(1,7): ws.cell(row=tr,column=c).fill=PatternFill("solid",fgColor=INK); ws.cell(row=tr,column=c).border=bd
    ws.cell(row=tr,column=3,value=f"=SUM(C{mfirst}:C{tr-1})").font=Font(bold=True,color=WHITE)
    oc=ws.cell(row=tr,column=4,value=f"=IFERROR(AVERAGE(D{mfirst}:D{tr-1}),0)"); oc.number_format=PCT; oc.font=Font(bold=True,color=WHITE)
    pp=ws.cell(row=tr,column=5,value=f"=SUM(E{mfirst}:E{tr-1})"); pp.number_format=SEK; pp.font=Font(bold=True,color=WHITE)
    ws.cell(row=tr,column=6,value=f"=SUM(F{mfirst}:F{tr-1})").font=Font(bold=True,color=WHITE)
    for i,w in enumerate([12,8,15,13,14,11],1): ws.column_dimensions[get_column_letter(i)].width=w

    # DASHBOARD
    ws=wb.create_sheet("Dashboard",0); ws.sheet_view.showGridLines=False
    ws["B2"]="Airbnb Studio — Dashboard"; ws["B2"].font=title_font
    ws["B3"]=f"Live from your account, {TODAY}."; ws["B3"].font=sub
    kpis=[("TOTAL ACTUAL PAYOUT",f"=SUM({PAY})",SEK),
          ("BOOKINGS PULLED",f"=COUNTA({ST})",'0'),
          ("NIGHTS BOOKED",f"=SUM('Monthly Summary'!$C${mfirst}:$C${tr-1})",'0'),
          ("AVG PAYOUT / BOOKING",f"=IFERROR(SUM({PAY})/COUNTIF({PAY},\"<>0\"),0)",SEK),
          ("AVG OCCUPANCY %",f"=IFERROR(AVERAGE('Monthly Summary'!$D${mfirst}:$D${tr-1}),0)",PCT),
          ("CONFIRMED + PENDING",f'=COUNTIF({ST},"Confirmed")+COUNTIF({ST},"Pending")','0')]
    pos=[("B","C",5),("E","F",5),("H","I",5),("B","C",9),("E","F",9),("H","I",9)]
    for (lbl_,v,fmt),(c1,c2,row) in zip(kpis,pos):
        ws.merge_cells(f"{c1}{row}:{c2}{row}"); ws.merge_cells(f"{c1}{row+1}:{c2}{row+1}")
        a=ws[f"{c1}{row}"]; a.value=lbl_; a.fill=PatternFill("solid",fgColor=TEAL); a.font=Font(bold=True,color=AQUA,size=10); a.alignment=ctr
        b=ws[f"{c1}{row+1}"]; b.value=v; b.fill=PatternFill("solid",fgColor=TEAL); b.font=Font(bold=True,color=WHITE,size=18); b.alignment=ctr; b.number_format=fmt
    ws["B12"]="Next check-in:"; ws["B12"].font=lbl
    ws["C12"]=f'=IFERROR(MIN(IF(Bookings!E{first}:E{last}="Confirmed",Bookings!B{first}:B{last})),"")'; ws["C12"].number_format=DATEF
    for col,w in [("A",3),("B",22),("C",16),("D",3),("E",18),("F",10),("G",3),("H",20),("I",10)]:
        ws.column_dimensions[col].width=w
    wb.save(MODEL)


def main():
    res=newest("reservations-*.json")
    if not res:
        sys.exit("No reservations-*.json in raw/. Run scrape_airbnb.py first.")
    rows=parse_reservations(res)
    if not rows:
        sys.exit("Parsed 0 bookings. Check raw/_debug/ — the table format may have changed.")
    (RAW/"bookings-normalized.json").write_text(json.dumps({"ingested":TODAY,"bookings":rows},ensure_ascii=False,indent=2),encoding="utf-8")
    # Clean CSV fact table for Power BI (stable path, typed columns, refresh-friendly).
    import csv
    with open(RAW/"bookings.csv","w",newline="",encoding="utf-8-sig") as f:
        w=csv.writer(f)
        w.writerow(["Guest","CheckIn","CheckOut","Nights","Status","ConfCode","Payout","Month"])
        for b in rows:
            w.writerow([b["guest"],b["check_in"],b["check_out"] or "",b["nights"] if b["nights"] is not None else "",
                        b["status"],b["confirmation"] or "",b["payout"] if b["payout"] is not None else "",b["check_in"][:7]])
    build(rows)
    active=[r for r in rows if r["status"] in ("Confirmed","Pending")]
    print(f"[ok] {len(rows)} bookings parsed ({len(active)} active/upcoming).")
    print(f"[ok] model -> {MODEL.relative_to(ROOT)}")
    print(f"[ok] normalized -> raw/bookings-normalized.json")


if __name__=="__main__":
    main()
